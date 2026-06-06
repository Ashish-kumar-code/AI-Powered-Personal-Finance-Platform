import os
import csv
from io import BytesIO
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from database import get_db_connection
from analytics_service import get_analytics_summary
from insights_service import generate_ai_insights

REPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "generated_reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

def get_report_data(user_id, month=None, year=None):
    """
    Combines analytics, transactions, and insights into a single package.
    """
    now = datetime.now()
    m = int(month) if month else now.month
    y = int(year) if year else now.year
    
    summary = get_analytics_summary(user_id, m, y)
    insights = generate_ai_insights(user_id)
    
    conn = get_db_connection()
    df_tx = pd.read_sql_query(
        "SELECT id, type, category, amount, date, description FROM transactions WHERE user_id = ? ORDER BY date DESC",
        conn,
        params=[user_id]
    )
    conn.close()
    
    # Filter transaction DataFrame to current month/year for detailed report sections
    df_tx['amount'] = pd.to_numeric(df_tx['amount'], errors='coerce').fillna(0.0)
    df_tx['date_parsed'] = pd.to_datetime(df_tx['date'])
    df_curr_tx = df_tx[(df_tx['date_parsed'].dt.month == m) & (df_tx['date_parsed'].dt.year == y)].copy()
    
    # Drop temp parsed column
    df_curr_tx = df_curr_tx.drop(columns=['date_parsed'])
    
    return {
        "month": m,
        "year": y,
        "summary": summary,
        "insights": insights,
        "transactions": df_curr_tx.to_dict('records')
    }

def generate_csv_report(user_id, month=None, year=None):
    """
    Generates a CSV string containing all transaction rows for the specified month/year.
    """
    data = get_report_data(user_id, month, year)
    output = BytesIO()
    
    # Create text wrapper
    wrapper = BytesIO()
    csv_writer = csv.writer(csv.UnixDialect())
    
    # Header
    rows = [
        ["AI-Powered Finance Ledger Report", f"Month: {data['month']}/{data['year']}"],
        [],
        ["Transaction ID", "Type", "Category", "Amount (INR)", "Date", "Description"]
    ]
    
    for tx in data["transactions"]:
        rows.append([tx["id"], tx["type"], tx["category"], tx["amount"], tx["date"], tx["description"]])
        
    # Write to memory stream
    csv_string = ""
    for r in rows:
        # csv writer expects a file-like object
        # we can just join or write to string
        csv_string += ",".join(f'"{str(val)}"' for val in r) + "\n"
        
    return csv_string.encode('utf-8')

def generate_excel_report(user_id, month=None, year=None):
    """
    Generates a professionally styled Excel file using openpyxl.
    """
    data = get_report_data(user_id, month, year)
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Financial Summary & KPIs ---
    ws1 = wb.active
    ws1.title = "Financial Summary"
    
    # Title Block
    ws1['A1'] = "AI-Powered Personal Finance Analytics Report"
    ws1['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws1['A2'] = f"Report Generation Period: {data['month']}/{data['year']}"
    ws1['A2'].font = Font(size=11, italic=True)
    
    # KPI Grid
    kpi_labels = [
        ("Total Income", data["summary"]["total_income"], "₹#,##0.00", "E2EFDA"), # Light green
        ("Total Expense", data["summary"]["total_expense"], "₹#,##0.00", "FCE4D6"), # Light orange
        ("Net Savings", data["summary"]["total_income"] - data["summary"]["total_expense"], "₹#,##0.00", "D9E1F2"), # Light blue
        ("Savings Rate", data["summary"]["savings_rate"] / 100.0, "0.0%", "FFF2CC"), # Light yellow
        ("Financial Health Score", data["summary"]["financial_health_score"], "0", "E2EFDA")
    ]
    
    # Write KPIs
    ws1['A4'] = "Key Performance Indicators"
    ws1['A4'].font = Font(size=12, bold=True)
    
    for i, (label, val, fmt, fill_color) in enumerate(kpi_labels):
        row = 5 + i
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell_val = ws1.cell(row=row, column=2, value=val)
        cell_val.number_format = fmt
        cell_val.font = Font(bold=True)
        cell_val.alignment = Alignment(horizontal="right")
        cell_val.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
    # AI Insights section on Sheet 1
    ws1.cell(row=12, column=1, value="AI-Generated Recommendations").font = Font(size=12, bold=True, color="1F4E79")
    ws1.cell(row=13, column=1, value="Category").font = Font(bold=True)
    ws1.cell(row=13, column=2, value="Insight Observation").font = Font(bold=True)
    
    for idx, ins in enumerate(data["insights"]):
        r_idx = 14 + idx
        ws1.cell(row=r_idx, column=1, value=ins["category"]).font = Font(bold=True)
        ws1.cell(row=r_idx, column=2, value=ins["content"])
        
    # Auto-adjust column widths
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 80
    
    # --- Sheet 2: Budgets Utilization ---
    ws2 = wb.create_sheet(title="Budget Planning")
    ws2['A1'] = "Category Budgets & Expense Utilization"
    ws2['A1'].font = Font(size=14, bold=True, color="1F4E79")
    
    headers2 = ["Category", "Budget Limit (INR)", "Actual Spent (INR)", "Remaining (INR)", "Utilization %"]
    for col_idx, h in enumerate(headers2):
        cell = ws2.cell(row=3, column=col_idx+1, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx, b in enumerate(data["summary"]["budget_utilization"]):
        row = 4 + r_idx
        ws2.cell(row=row, column=1, value=b["category"])
        
        c_bud = ws2.cell(row=row, column=2, value=b["budget"])
        c_bud.number_format = "₹#,##0.00"
        
        c_spent = ws2.cell(row=row, column=3, value=b["spent"])
        c_spent.number_format = "₹#,##0.00"
        
        c_rem = ws2.cell(row=row, column=4, value=b["remaining"])
        c_rem.number_format = "₹#,##0.00"
        if b["remaining"] < 0:
            c_rem.font = Font(color="FF0000")  # red text for exceeded
            
        c_util = ws2.cell(row=row, column=5, value=b["utilization_pct"] / 100.0)
        c_util.number_format = "0.0%"
        if b["utilization_pct"] > 100:
            c_util.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red warning fill
            
    # Auto adjust column dimensions
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    # --- Sheet 3: Transactions Ledger ---
    ws3 = wb.create_sheet(title="Transaction Ledger")
    ws3['A1'] = "Detailed Transactions Register"
    ws3['A1'].font = Font(size=14, bold=True, color="1F4E79")
    
    headers3 = ["ID", "Type", "Category", "Amount (INR)", "Date", "Description"]
    for col_idx, h in enumerate(headers3):
        cell = ws3.cell(row=3, column=col_idx+1, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx, tx in enumerate(data["transactions"]):
        row = 4 + r_idx
        ws3.cell(row=row, column=1, value=tx["id"]).alignment = Alignment(horizontal="center")
        ws3.cell(row=row, column=2, value=tx["type"]).alignment = Alignment(horizontal="center")
        ws3.cell(row=row, column=3, value=tx["category"])
        
        c_amt = ws3.cell(row=row, column=4, value=tx["amount"])
        c_amt.number_format = "₹#,##0.00"
        if tx["type"] == "Income":
            c_amt.font = Font(color="375623")  # green color for income
        else:
            c_amt.font = Font(color="C65911")  # orange color for expense
            
        ws3.cell(row=row, column=5, value=tx["date"]).alignment = Alignment(horizontal="center")
        ws3.cell(row=row, column=6, value=tx["description"])
        
    # Column sizing
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save to stream
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream.getvalue()

def generate_pdf_report(user_id, month=None, year=None):
    """
    Generates a beautiful PDF report using ReportLab.
    """
    data = get_report_data(user_id, month, year)
    pdf_buffer = BytesIO()
    
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#555555"),
        spaceAfter=20
    )
    
    heading_style = ParagraphStyle(
        'SecHeading',
        parent=styles['Heading2'],
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#1F4E79"),
        spaceBefore=15,
        spaceAfter=10
    )
    
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=9,
        leading=11
    )
    
    bold_cell_style = ParagraphStyle(
        'BoldCell',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        fontName='Helvetica-Bold'
    )
    
    story = []
    
    # 1. Header
    story.append(Paragraph("AI-Powered Personal Finance Analytics Report", title_style))
    story.append(Paragraph(f"Reporting Cycle: {data['month']}/{data['year']} | Created: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    
    # 2. Section: KPIs & Health Score
    story.append(Paragraph("I. Key Metrics & Financial Health", heading_style))
    
    summary = data["summary"]
    kpi_table_data = [
        [Paragraph("Metric Title", bold_cell_style), Paragraph("Current Period Value", bold_cell_style)],
        [Paragraph("Total Income", table_cell_style), Paragraph(f"₹{summary['total_income']:.2f}", table_cell_style)],
        [Paragraph("Total Expenses", table_cell_style), Paragraph(f"₹{summary['total_expense']:.2f}", table_cell_style)],
        [Paragraph("Net Savings Summary", table_cell_style), Paragraph(f"₹{(summary['total_income'] - summary['total_expense']):.2f}", table_cell_style)],
        [Paragraph("Calculated Savings Rate", table_cell_style), Paragraph(f"{summary['savings_rate']}%", table_cell_style)],
        [Paragraph("Algorithmic Health Score", bold_cell_style), Paragraph(f"<b>{summary['financial_health_score']} / 100</b>", bold_cell_style)]
    ]
    
    kpi_table = Table(kpi_table_data, colWidths=[200, 200])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (1,0), colors.HexColor("#1F4E79")),
        ('TEXTCOLOR', (0,0), (1,0), colors.white),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D3D3D3")),
        ('BACKGROUND', (0,1), (-1,-2), colors.HexColor("#F9F9F9")),
        ('BACKGROUND', (0,-1), (1,-1), colors.HexColor("#E2EFDA")), # Light green highlights for score
    ]))
    
    # Fix white text colors for header row inside Paragraphs:
    kpi_table_data[0][0].style.textColor = colors.white
    kpi_table_data[0][1].style.textColor = colors.white
    
    story.append(kpi_table)
    story.append(Spacer(1, 15))
    
    # 3. Section: Budget Utilizations
    story.append(Paragraph("II. Category Budgets & Expenses Comparison", heading_style))
    
    budget_headers = [
        Paragraph("Category", bold_cell_style),
        Paragraph("Monthly Limit", bold_cell_style),
        Paragraph("Actual Spent", bold_cell_style),
        Paragraph("Remaining Balance", bold_cell_style),
        Paragraph("Usage %", bold_cell_style)
    ]
    
    budget_headers[0].style.textColor = colors.white
    budget_headers[1].style.textColor = colors.white
    budget_headers[2].style.textColor = colors.white
    budget_headers[3].style.textColor = colors.white
    budget_headers[4].style.textColor = colors.white
    
    budget_table_data = [budget_headers]
    
    for b in summary["budget_utilization"]:
        rem_color = "red" if b["remaining"] < 0 else "black"
        rem_text = f"<font color='{rem_color}'>₹{b['remaining']:.2f}</font>"
        
        util_color = "red" if b["utilization_pct"] > 100 else "black"
        util_text = f"<font color='{util_color}'>{b['utilization_pct']:.1f}%</font>"
        
        budget_table_data.append([
            Paragraph(b["category"], table_cell_style),
            Paragraph(f"₹{b['budget']:.2f}", table_cell_style),
            Paragraph(f"₹{b['spent']:.2f}", table_cell_style),
            Paragraph(rem_text, table_cell_style),
            Paragraph(util_text, table_cell_style),
        ])
        
    if len(budget_table_data) == 1:
        budget_table_data.append([Paragraph("No budgets configured for this billing period.", table_cell_style), "", "", "", ""])
        
    budget_table = Table(budget_table_data, colWidths=[110, 100, 100, 110, 80])
    budget_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E79")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D3D3D3")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F9F9F9")]),
    ]))
    
    story.append(budget_table)
    story.append(Spacer(1, 15))
    
    # 4. Section: AI Recommendations
    story.append(Paragraph("III. AI Advisor Financial Analysis & Recommendations", heading_style))
    
    for ins in data["insights"]:
        bullet = "• "
        text = f"<b>[{ins['category']}] {ins['title']}:</b> {ins['content']}"
        story.append(Paragraph(bullet + text, table_cell_style))
        story.append(Spacer(1, 4))
        
    story.append(Spacer(1, 15))
    
    # 5. Section: Recent Transactions
    story.append(Paragraph("IV. Transactions Ledger (Current Period Limit: 15)", heading_style))
    
    tx_headers = [
        Paragraph("ID", bold_cell_style),
        Paragraph("Type", bold_cell_style),
        Paragraph("Category", bold_cell_style),
        Paragraph("Amount", bold_cell_style),
        Paragraph("Date", bold_cell_style),
        Paragraph("Description", bold_cell_style)
    ]
    tx_headers[0].style.textColor = colors.white
    tx_headers[1].style.textColor = colors.white
    tx_headers[2].style.textColor = colors.white
    tx_headers[3].style.textColor = colors.white
    tx_headers[4].style.textColor = colors.white
    tx_headers[5].style.textColor = colors.white
    
    tx_table_data = [tx_headers]
    
    # Show last 15
    for tx in data["transactions"][:15]:
        amt_color = "green" if tx["type"] == "Income" else "orange"
        amt_text = f"<font color='{amt_color}'>₹{tx['amount']:.2f}</font>"
        
        tx_table_data.append([
            Paragraph(str(tx["id"]), table_cell_style),
            Paragraph(tx["type"], table_cell_style),
            Paragraph(tx["category"], table_cell_style),
            Paragraph(amt_text, table_cell_style),
            Paragraph(tx["date"], table_cell_style),
            Paragraph(tx["description"] or "", table_cell_style),
        ])
        
    if len(tx_table_data) == 1:
        tx_table_data.append([Paragraph("No transactions logged for this month.", table_cell_style), "", "", "", "", ""])
        
    tx_table = Table(tx_table_data, colWidths=[40, 60, 90, 80, 80, 150])
    tx_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E79")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D3D3D3")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F9F9F9")]),
    ]))
    
    story.append(tx_table)
    
    # Build PDF
    doc.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()
